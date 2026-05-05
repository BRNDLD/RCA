from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class Options:
    origins: list[str]
    systems: list[str]
    action_types: list[str]
    yes_no: list[str]
    categories: dict[str, list[str]]

    def to_dict(self) -> dict[str, Any]:
        return {
            "origins": self.origins,
            "systems": self.systems,
            "actionTypes": self.action_types,
            "yesNo": self.yes_no,
            "categories": self.categories,
        }


_cached: Options | None = None
_cached_path: str | None = None


def get_options(excel_path: str) -> Options:
    global _cached, _cached_path

    if _cached is not None and _cached_path == excel_path:
        return _cached

    path = Path(excel_path)
    if not path.exists():
        # Return safe fallbacks if file isn't present.
        opts = Options(
            origins=[],
            systems=[],
            action_types=["Correctiva", "Mejora"],
            yes_no=["SÍ", "NO"],
            categories={},
        )
        _cached = opts
        _cached_path = excel_path
        return opts

    wb = load_workbook(path, data_only=True)
    if "Datos" not in wb.sheetnames:
        opts = Options(
            origins=[],
            systems=[],
            action_types=["Correctiva", "Mejora"],
            yes_no=["SÍ", "NO"],
            categories={},
        )
        _cached = opts
        _cached_path = excel_path
        return opts

    ws = wb["Datos"]

    def _norm(v: object | None) -> str:
        if v is None:
            return ""
        return str(v).strip().lower()

    def _dedupe_preserve_order(values: list[str]) -> list[str]:
        seen: set[str] = set()
        uniq: list[str] = []
        for s in values:
            if s in seen:
                continue
            seen.add(s)
            uniq.append(s)
        return uniq

    def _read_down_block(
        col: int, start_row: int, *, max_rows: int = 200
    ) -> tuple[list[str], int]:
        """Reads a vertical list until first blank cell.

        Returns (unique_values, first_blank_row).
        """
        out: list[str] = []
        stop_row = min(start_row + max_rows, ws.max_row + 1)
        for r in range(start_row, stop_row):
            v = ws.cell(r, col).value
            if v is None or str(v).strip() == "":
                return (_dedupe_preserve_order(out), r)
            out.append(str(v).strip())
        return (_dedupe_preserve_order(out), stop_row)

    def _read_down_range(col: int, start_row: int, end_row_inclusive: int) -> list[str]:
        out: list[str] = []
        for r in range(start_row, min(end_row_inclusive, ws.max_row) + 1):
            v = ws.cell(r, col).value
            if v is None or str(v).strip() == "":
                break
            out.append(str(v).strip())
        return _dedupe_preserve_order(out)

    def _find_cell(target: str, *, max_row: int = 20, max_col: int = 20) -> tuple[int, int] | None:
        t = target.strip().lower()
        for r in range(1, min(max_row, ws.max_row) + 1):
            for c in range(1, min(max_col, ws.max_column) + 1):
                if _norm(ws.cell(r, c).value) == t:
                    return (r, c)
        return None

    # Origen No conformidad block
    origins: list[str] = []
    block_end_row: int | None = None
    origin_hdr = _find_cell("Origen No conformidad", max_row=10, max_col=10)
    if origin_hdr:
        r, c = origin_hdr
        origins, block_end_row = _read_down_block(c, r + 1)

    # Systems block: first item is typically "SG" (same row as first origin)
    systems: list[str] = []
    sg_cell = _find_cell("SG", max_row=15, max_col=10)
    if sg_cell:
        r, c = sg_cell
        if block_end_row is not None:
            systems = _read_down_range(c, r, block_end_row - 1)
        else:
            systems, _ = _read_down_block(c, r)

    # Type of Action block
    action_types: list[str] = []
    toa_hdr = _find_cell("Type of Action", max_row=15, max_col=12)
    if toa_hdr:
        r, c = toa_hdr
        action_types, _ = _read_down_block(c, r + 1)

    # Yes/No block (starts at first "SÍ")
    yes_no: list[str] = []
    si_cell = _find_cell("SÍ", max_row=25, max_col=12)
    if si_cell:
        r, c = si_cell
        yes_no, _ = _read_down_block(c, r)

    # Categories block: headers appear in the same row as "Proceso"
    categories: dict[str, list[str]] = {}
    proceso_cell = _find_cell("Proceso", max_row=25, max_col=20)
    if proceso_cell:
        header_row, _ = proceso_cell
        for label in ["Proceso", "Producto", "Negocio", "Mano de Obra"]:
            loc = None
            # Search within the detected header row to avoid picking later blocks.
            for c in range(1, ws.max_column + 1):
                if _norm(ws.cell(header_row, c).value) == label.lower():
                    loc = (header_row, c)
                    break
            if not loc:
                continue
            _, col = loc
            if block_end_row is not None:
                categories[label] = _read_down_range(col, header_row + 1, block_end_row - 1)
            else:
                categories[label], _ = _read_down_block(col, header_row + 1)

    opts = Options(
        origins=origins,
        systems=systems,
        action_types=action_types,
        yes_no=yes_no,
        categories=categories,
    )
    _cached = opts
    _cached_path = excel_path
    return opts
